from flask import Flask, render_template, request, jsonify
import subprocess
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start_monitoring', methods=['POST'])
def start_monitoring():
    subprocess.call(["python", "sbm.py"])
    return jsonify(status="Monitoring started")

@app.route('/get_average_bpm', methods=['GET'])
def get_average_bpm():
    wb = openpyxl.load_workbook("datas/heart_rate_data.xlsx")
    ws = wb.active
    average_bpm = ws.cell(row=ws.max_row, column=2).value
    return jsonify(average_bpm=average_bpm)

if __name__ == '__main__':
    app.run(debug=True)
